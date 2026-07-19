import csv
from dataclasses import dataclass, replace
from pathlib import Path

from openpyxl import load_workbook

from .models import ProfileMetrics


_REQUIRED_COLUMNS = {
    "commenter_user_id",
    "commenter_handle",
    "commenter_profile_url",
    "video_id",
}

_REQUIRED_USER_EXPORT_COLUMNS = {
    "username",
    "user_id",
    "sec_uid",
    "private",
    "follower_count",
    "following_count",
    "video_count",
}

_REQUIRED_WORKBOOK_COLUMNS = {
    "账号",
    "secUid",
    "用户ID",
    "私密账号",
    "该粉丝粉丝数",
    "该粉丝关注数",
    "视频数",
    "主页链接",
}


@dataclass(frozen=True)
class Target:
    target_id: str
    username: str
    profile_url: str
    source_video_id: str
    sec_uid: str
    profile_metrics: ProfileMetrics | None = None
    private_account: bool | None = None
    identity_key: str = ""
    source_line_numbers: tuple[int, ...] = ()


@dataclass(frozen=True)
class ImportResult:
    targets: tuple[Target, ...]
    skipped_duplicates: int
    skipped_invalid: int = 0


def target_identity_key(*, sec_uid: str, target_id: str, username: str) -> str:
    if normalized := sec_uid.strip():
        return f"sec:{normalized}"
    if normalized := target_id.strip():
        return f"uid:{normalized}"
    if normalized := username.strip().removeprefix("@").lower():
        return f"handle:{normalized}"
    raise ValueError("target identity is empty")


def read_targets(path: Path) -> ImportResult:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_follower_workbook(path)
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        columns = set(next(csv.reader(source), ()))
    if _REQUIRED_USER_EXPORT_COLUMNS <= columns:
        return _read_deduplicated_user_export(path)
    return _read_comment_export(path)


def _read_deduplicated_user_export(path: Path) -> ImportResult:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        targets: list[Target] = []
        target_indexes: dict[str, int] = {}
        skipped_duplicates = 0
        skipped_invalid = 0
        for line_number, row in enumerate(reader, start=2):
            target_id = str(row.get("user_id") or "").strip()
            username = str(row.get("username") or "").strip().removeprefix("@").lower()
            sec_uid = str(row.get("sec_uid") or "").strip()
            if not target_id or not username:
                skipped_invalid += 1
                continue
            identity_key = target_identity_key(
                sec_uid=sec_uid,
                target_id=target_id,
                username=username,
            )
            if identity_key in target_indexes:
                index = target_indexes[identity_key]
                existing = targets[index]
                targets[index] = replace(
                    existing,
                    source_line_numbers=existing.source_line_numbers + (line_number,),
                )
                skipped_duplicates += 1
                continue
            target_indexes[identity_key] = len(targets)
            targets.append(
                Target(
                    target_id=target_id,
                    username=username,
                    profile_url=f"https://www.tiktok.com/@{username}",
                    source_video_id="",
                    sec_uid=sec_uid,
                    profile_metrics=ProfileMetrics(
                        following=_csv_int(row, "following_count"),
                        followers=_csv_int(row, "follower_count"),
                        posts=_csv_int(row, "video_count"),
                    ),
                    private_account=str(row.get("private") or "").strip().lower()
                    in {"1", "true", "yes", "是"},
                    identity_key=identity_key,
                    source_line_numbers=(line_number,),
                )
            )
    return ImportResult(
        targets=tuple(targets),
        skipped_duplicates=skipped_duplicates,
        skipped_invalid=skipped_invalid,
    )


def _csv_int(row: dict[str, str], field: str) -> int:
    raw = str(row.get(field) or "").strip().replace(",", "")
    try:
        return max(0, int(float(raw))) if raw else 0
    except ValueError:
        return 0


def _read_comment_export(path: Path) -> ImportResult:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        columns = set(reader.fieldnames or ())
        missing = sorted(_REQUIRED_COLUMNS - columns)
        if missing:
            raise ValueError(f"missing required CSV columns: {', '.join(missing)}")

        targets: list[Target] = []
        target_indexes: dict[str, int] = {}
        skipped_duplicates = 0
        for line_number, row in enumerate(reader, start=2):
            target_id = (row["commenter_user_id"] or "").strip()
            username = (row["commenter_handle"] or "").strip().removeprefix("@").lower()
            if not target_id or not username:
                raise ValueError(f"missing target identity on CSV line {line_number}")
            sec_uid = (row.get("commenter_sec_uid") or "").strip()
            identity_key = target_identity_key(
                sec_uid=sec_uid,
                target_id=target_id,
                username=username,
            )
            if identity_key in target_indexes:
                index = target_indexes[identity_key]
                existing = targets[index]
                targets[index] = replace(
                    existing,
                    source_line_numbers=existing.source_line_numbers + (line_number,),
                )
                skipped_duplicates += 1
                continue
            target_indexes[identity_key] = len(targets)
            targets.append(
                Target(
                    target_id=target_id,
                    username=username,
                    profile_url=(row["commenter_profile_url"] or "").strip(),
                    source_video_id=(row["video_id"] or "").strip(),
                    sec_uid=sec_uid,
                    identity_key=identity_key,
                    source_line_numbers=(line_number,),
                )
            )

    return ImportResult(targets=tuple(targets), skipped_duplicates=skipped_duplicates)


def _read_follower_workbook(path: Path) -> ImportResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(str(value or "").strip() for value in next(rows, ()))
        columns = {name: index for index, name in enumerate(headers) if name}
        missing = sorted(_REQUIRED_WORKBOOK_COLUMNS - set(columns))
        if missing:
            raise ValueError(f"missing required workbook columns: {', '.join(missing)}")

        targets: list[Target] = []
        seen_ids: set[str] = set()
        skipped_duplicates = 0
        skipped_invalid = 0
        for row in rows:
            target_id = _cell_text(row, columns["用户ID"])
            username = _cell_text(row, columns["账号"]).removeprefix("@").lower()
            if not target_id or not username:
                skipped_invalid += 1
                continue
            if target_id in seen_ids:
                skipped_duplicates += 1
                continue
            seen_ids.add(target_id)
            targets.append(
                Target(
                    target_id=target_id,
                    username=username,
                    profile_url=_cell_text(row, columns["主页链接"]),
                    source_video_id="",
                    sec_uid=_cell_text(row, columns["secUid"]),
                    profile_metrics=ProfileMetrics(
                        following=_cell_int(row, columns["该粉丝关注数"]),
                        followers=_cell_int(row, columns["该粉丝粉丝数"]),
                        posts=_cell_int(row, columns["视频数"]),
                    ),
                    private_account=_cell_bool(row, columns["私密账号"]),
                )
            )
        return ImportResult(
            targets=tuple(targets),
            skipped_duplicates=skipped_duplicates,
            skipped_invalid=skipped_invalid,
        )
    finally:
        workbook.close()


def _cell_text(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _cell_int(row: tuple[object, ...], index: int) -> int:
    raw = _cell_text(row, index).replace(",", "")
    if not raw:
        return 0
    try:
        return max(0, int(float(raw)))
    except ValueError:
        return 0


def _cell_bool(row: tuple[object, ...], index: int) -> bool:
    return _cell_text(row, index).lower() in {"1", "true", "yes", "是"}
