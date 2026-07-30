import json
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from openpyxl import load_workbook

from .importer import Target, target_identity_key

_REQUIRED_PRIORITY_WORKBOOK_COLUMNS = {
    "follower_handle",
    "follower_uid",
    "follower_sec_uid",
}
_HANDLE_PATTERN = re.compile(r"[A-Za-z0-9._]{1,24}")
_JSON_TEXT_FIELDS = (
    "username",
    "user_id",
    "sec_uid",
    "profile_url",
    "source_video_id",
    "source_live_id",
    "source_type",
    "source_id",
    "collected_at",
)
_QUALITY_REASONS = {
    "comment",
    "follow",
    "gift",
    "share",
    "like",
    "multiple_event_types",
    "multiple_rooms",
}


@dataclass(frozen=True)
class PriorityImportResult:
    targets: tuple[Target, ...]
    source_rows: int
    skipped_duplicates: int
    skipped_invalid: int


def read_priority_targets(path: Path, *, source_live_id: str) -> PriorityImportResult:
    source = Path(path)
    live_id = str(source_live_id).strip()
    if not live_id:
        raise ValueError("source_live_id is required")
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_workbook(source)
    if source.suffix.lower() not in {".jsonl", ".ndjson"}:
        raise ValueError("priority input must be JSONL, XLSX, or XLSM")
    return _read_jsonl(source, live_id)


def _read_jsonl(path: Path, live_id: str) -> PriorityImportResult:
    targets: list[Target] = []
    aliases: dict[str, int] = {}
    source_rows = 0
    skipped_duplicates = 0
    skipped_invalid = 0
    with path.open("r", encoding="utf-8-sig") as source:
        for line_number, raw in enumerate(source, start=1):
            if not raw.strip():
                continue
            source_rows += 1
            try:
                row = json.loads(raw)
            except json.JSONDecodeError as error:
                raise ValueError(f"JSONL line {line_number} is malformed") from error
            if not isinstance(row, dict):
                raise TypeError(f"JSONL line {line_number} must be an object")
            _validate_quality_metadata(row, line_number)
            values = {
                field: _json_text(row, field, line_number)
                for field in _JSON_TEXT_FIELDS
            }
            source_type = str(values["source_type"] or "").strip()
            if source_type and source_type not in {
                "followers",
                "comments",
                "live",
                "live_audience",
            }:
                raise ValueError(
                    f"JSONL line {line_number} has an unsupported source_type"
                )
            row_source_id = str(
                values["source_id"] or values["source_live_id"] or live_id
            ).strip()
            if row_source_id != live_id:
                field = "source_id" if values["source_id"] else "source_live_id"
                raise ValueError(f"JSONL line {line_number} has a different {field}")
            target = _target_from_values(
                username=values["username"],
                target_id=values["user_id"],
                sec_uid=values["sec_uid"],
                profile_url=values["profile_url"],
                source_video_id=values["source_video_id"],
                source_line_number=line_number,
            )
            if target is None:
                skipped_invalid += 1
                continue
            skipped_duplicates += _merge_target(targets, aliases, target)
    return PriorityImportResult(
        tuple(targets), source_rows, skipped_duplicates, skipped_invalid
    )


def _read_workbook(path: Path) -> PriorityImportResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(str(value or "").strip() for value in next(rows, ()))
        columns = {name: index for index, name in enumerate(headers) if name}
        missing = sorted(_REQUIRED_PRIORITY_WORKBOOK_COLUMNS - set(columns))
        if missing:
            raise ValueError(
                "missing required priority workbook columns: " + ", ".join(missing)
            )
        targets: list[Target] = []
        aliases: dict[str, int] = {}
        source_rows = 0
        skipped_duplicates = 0
        skipped_invalid = 0
        for line_number, row in enumerate(rows, start=2):
            source_rows += 1
            target = _target_from_values(
                username=_cell(row, columns["follower_handle"]),
                target_id=_cell(row, columns["follower_uid"]),
                sec_uid=_cell(row, columns["follower_sec_uid"]),
                profile_url=None,
                source_video_id=None,
                source_line_number=line_number,
            )
            if target is None:
                skipped_invalid += 1
                continue
            skipped_duplicates += _merge_target(targets, aliases, target)
        return PriorityImportResult(
            tuple(targets), source_rows, skipped_duplicates, skipped_invalid
        )
    finally:
        workbook.close()


def _target_from_values(
    *,
    username: Any,
    target_id: Any,
    sec_uid: Any,
    profile_url: Any,
    source_video_id: Any,
    source_line_number: int,
) -> Target | None:
    normalized_username = str(username or "").strip().removeprefix("@").lower()
    if _HANDLE_PATTERN.fullmatch(normalized_username) is None:
        return None
    raw_target_id = str(target_id or "").strip()
    normalized_target_id = (
        "" if raw_target_id.lower().startswith("dom-") else raw_target_id
    )
    normalized_sec_uid = str(sec_uid or "").strip()
    normalized_profile_url = _profile_url(
        profile_url, normalized_username, source_line_number
    )
    identity_key = target_identity_key(
        sec_uid=normalized_sec_uid,
        target_id=normalized_target_id,
        username=normalized_username,
    )
    return Target(
        target_id=normalized_target_id,
        username=normalized_username,
        profile_url=normalized_profile_url,
        source_video_id=str(source_video_id or "").strip(),
        sec_uid=normalized_sec_uid,
        identity_key=identity_key,
        source_line_numbers=(source_line_number,),
    )


def _merge_target(
    targets: list[Target], aliases: dict[str, int], incoming: Target
) -> int:
    keys = _target_aliases(incoming)
    matching = {aliases[key] for key in keys if key in aliases}
    if not matching:
        index = len(targets)
        targets.append(incoming)
        for key in keys:
            aliases[key] = index
        return 0

    index = min(matching)
    merged = targets[index]
    for other_index in sorted(matching - {index}, reverse=True):
        merged = _combine_targets(merged, targets[other_index])
        del targets[other_index]
        for key, value in tuple(aliases.items()):
            if value == other_index:
                aliases[key] = index
            elif value > other_index:
                aliases[key] = value - 1
    merged = _combine_targets(merged, incoming)
    targets[index] = merged
    for key in _target_aliases(merged):
        aliases[key] = index
    return len(matching)


def _combine_targets(first: Target, second: Target) -> Target:
    line_number = second.source_line_numbers[0]
    if first.sec_uid and second.sec_uid and first.sec_uid != second.sec_uid:
        raise ValueError(f"source line {line_number} has conflicting sec_uid")
    if first.target_id and second.target_id and first.target_id != second.target_id:
        raise ValueError(f"source line {line_number} has conflicting user_id")
    return replace(
        first,
        target_id=second.target_id or first.target_id,
        username=second.username or first.username,
        profile_url=second.profile_url or first.profile_url,
        source_video_id=second.source_video_id or first.source_video_id,
        sec_uid=second.sec_uid or first.sec_uid,
        identity_key=target_identity_key(
            sec_uid=second.sec_uid or first.sec_uid,
            target_id=second.target_id or first.target_id,
            username=second.username or first.username,
        ),
        source_line_numbers=first.source_line_numbers + second.source_line_numbers,
    )


def _target_aliases(target: Target) -> tuple[str, ...]:
    aliases = [f"handle:{target.username}"]
    if target.sec_uid:
        aliases.append(f"sec:{target.sec_uid}")
    if target.target_id:
        aliases.append(f"uid:{target.target_id}")
    return tuple(aliases)


def _cell(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _json_text(row: dict[str, Any], field: str, line_number: int) -> str | None:
    value = row.get(field)
    if value is None or isinstance(value, str):
        return value
    raise ValueError(f"JSONL line {line_number} field {field} must be a string or null")


def _validate_quality_metadata(row: dict[str, Any], line_number: int) -> None:
    has_level = "lead_level" in row
    has_reasons = "qualification_reasons" in row
    if not has_level and not has_reasons:
        return
    if not has_level:
        raise ValueError(f"JSONL line {line_number} lead_level is required")
    if not has_reasons:
        raise ValueError(f"JSONL line {line_number} qualification_reasons is required")

    level = row["lead_level"]
    if level not in {"A", "B"}:
        raise ValueError(f"JSONL line {line_number} lead_level must be A or B")
    reasons = row["qualification_reasons"]
    if not isinstance(reasons, list):
        raise TypeError(
            f"JSONL line {line_number} qualification_reasons must be an array"
        )
    if not reasons:
        raise ValueError(
            f"JSONL line {line_number} qualification_reasons must be nonempty"
        )
    if any(not isinstance(reason, str) or not reason.strip() for reason in reasons):
        raise ValueError(
            f"JSONL line {line_number} qualification_reasons must contain "
            "nonempty strings"
        )
    if any(reason not in _QUALITY_REASONS for reason in reasons):
        raise ValueError(
            f"JSONL line {line_number} qualification_reasons contains an "
            "unsupported value"
        )


def _profile_url(value: Any, username: str, line_number: int) -> str:
    canonical = f"https://www.tiktok.com/@{username}"
    raw = str(value or "").strip()
    if not raw:
        return canonical
    parsed = urlsplit(raw)
    if (
        parsed.scheme.lower() != "https"
        or parsed.hostname not in {"tiktok.com", "www.tiktok.com"}
        or parsed.path.rstrip("/").lower() != f"/@{username}".lower()
        or parsed.query
        or parsed.fragment
    ):
        raise ValueError(
            f"source line {line_number} profile_url must be a TikTok profile URL"
        )
    return canonical
